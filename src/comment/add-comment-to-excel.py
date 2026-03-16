import json
from openpyxl import load_workbook
from openpyxl.comments import Comment
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
input_file = ROOT / "output" / "stock-charts.xlsx"
output_file = ROOT / "output" / "stock-charts-with-comments.xlsx"
data_file = Path(__file__).parent / "data" / "comments.json"

wb = load_workbook(input_file)
ws = wb.active

with open(data_file, 'r', encoding='utf-8') as f:
    summary_list = json.load(f)

summary_dict = {item['code']: item['summary'] for item in summary_list}

for row in ws.iter_rows(min_row=2, max_col=2):
    stock_code = row[1].value

    if stock_code in summary_dict:
        summary_text = summary_dict[stock_code]
        comment = Comment(summary_text, "Bot")
        comment.width = 300
        line_count = summary_text.count('\n') + 1
        comment.height = max(50, line_count * 18)
        row[0].comment = comment

wb.save(output_file)
print("comments added")
